from tkinter import *
import os
from tkinter import messagebox
import openpyxl
from PIL import Image, ImageDraw, ImageFont

font_class = ImageFont.truetype('Noah Bold.ttf', size= 14 * 3)
font_clMain = ImageFont.truetype('Noah Bold.ttf', size= 25 * 3)
font_lesson = ImageFont.truetype('Noah Medium.ttf', size= 15 * 3)
font_teacher = ImageFont.truetype('Noah Medium.ttf', size= 13 * 3)

window = Tk()
window.title("Мастер создания расписания")
window.geometry('700x470')
window.resizable(width=False, height=False)
window["bg"] = "AntiqueWhite1"
window["padx"] = "30"
window["pady"] = "30"


def func_make(ws, clas, j):
    sample = Image.open('Raspisanie-sample.png')


    def f(text):
        if str(text) == 'None':
            return ''
        else:
            return text
    draw_text = ImageDraw.Draw(sample)
    for m, mMax, count in zip(range(0, 1700, 1120), range(0, 20, 19), range(1, 3)):
        listX = [615, 1082, 1550, 2017, 2485, 2953]
        listY = [319, 445, 571, 697, 823, 949, 1075]
        draw_text.text((120, 140 + m), clas[:-1] + str(count), font=font_clMain, fill=('#000000'))
        for x, xChar in zip(range(6), range(67, 78, 2)):
            for y, yInt in zip(range(7), range(5, 18, 2)):
                lesson = ws[f'{chr(xChar)}{yInt + mMax + j -1}'].value
                room = ws[f'{chr(xChar+1)}{yInt + mMax +  j -1}'].value
                teacher = ws[f'{chr(xChar)}{yInt + 1 + mMax + j -1}'].value
                draw_text.text((listX[x], listY[y] + m), f(lesson), font=font_lesson, fill=('#000000'))
                draw_text.text((listX[x]+360, listY[y]+6 + m), f(room), font=font_class, fill=('#c5c3c3'))
                draw_text.text((listX[x], listY[y] + 55 + m), f(teacher), font=font_teacher, fill=('#c5c3c3'))

    print(f'Расписание готово {clas}')
    try:
        sample.save(pngway + f"\\{clas[:-2]} класс\\{clas[:-1]}.png")
    except:
        os.mkdir(pngway + f"\\{clas[:-2]} класс")
        sample.save(pngway + f"\\{clas[:-2]} класс\\{clas[:-1]}.png")


def find_clas():
    table = openpyxl.open(xlsway, read_only=True)
    sheet = table.sheetnames
    for i in range(0, 4):
        count = 0
        for j in range(1, 211, 19):
            count += 1
            ws = table[sheet[i]]
            text = ws[f'A{j}'].value
            print(text)
            if str(text) != 'None' and count % 2 == 1:
                func_make(ws, text[-4:], j)
    messagebox.showinfo('Готово!', f'Расписание готово! \nПапка: {pngway}')


def clicked1():
    global xlsway
    res = "Выбранная папка:\n{}".format(textInput1.get())
    xlsway = "{}".format(textInput1.get())
    xls2Label1.configure(text=res)
    print(xlsway)


def clicked2():
    global pngway
    res = "Выбранная папка:\n{}".format(textInput2.get())
    pngway = "{}".format(textInput2.get())
    xls2Label2.configure(text=res)
    print(pngway)


topLabel = Label(window,
                 text="Мастер создания расписания",
                 font=("Noah ExtraBold", 24),
                 bg="AntiqueWhite1")
topLabel.pack()

xlsLabel1 = Label(window,
                  text="Укажите путь к файлу эксель расписания",
                  font=("Noah Regular", 15),
                  bg="AntiqueWhite1")
xlsLabel1.pack()
xlsFrame1 = Frame(window,
                  padx=5,
                  pady=5,
                  bg="AntiqueWhite1")
xlsFrame1.pack()
textInput1 = Entry(xlsFrame1,
                   font=("Noah Regular", 9),
                   width=50,
                   bg="AntiqueWhite1")
textInput1.grid(column=0, row=0)
Button1 = Button(xlsFrame1,
                 text="Сохранить",
                 bg="AntiqueWhite2",
                 fg="AntiqueWhite4",
                 activebackground='AntiqueWhite3',
                 command=clicked1)
Button1.grid(column=1, row=0)

xls2Label1 = Label(window,
                   text="C:\ ",
                   font=("Noah Regular Italic", 12),
                   bg="AntiqueWhite1",
                   fg="AntiqueWhite4")
xls2Label1.pack()

xlsLabel2 = Label(window,
                  text="Укажите папку, куда сохранить картинки",
                  font=("Noah Regular", 15),
                  bg="AntiqueWhite1")
xlsLabel2.pack()
xlsFrame2 = Frame(window,
                  padx=5,
                  pady=5,
                  bg="AntiqueWhite1")
xlsFrame2.pack()
textInput2 = Entry(xlsFrame2,
                   font=("Noah Regular", 9),
                   width=50,
                   bg="AntiqueWhite1")
textInput2.grid(column=0, row=0)
Button2 = Button(xlsFrame2,
                 text="Сохранить",
                 bg="AntiqueWhite2",
                 fg="AntiqueWhite4",
                 activebackground='AntiqueWhite3',
                 command=clicked2)
Button2.grid(column=1, row=0)

xls2Label2 = Label(window,
                   text="C:\ ",
                   font=("Noah Regular Italic", 12),
                   bg="AntiqueWhite1",
                   fg="AntiqueWhite4")
xls2Label2.pack()

Button3 = Button(window,
                 height="2",
                 width="20",
                 text="Создать расписание!",
                 bg="AntiqueWhite2",
                 fg="AntiqueWhite4",
                 activebackground='AntiqueWhite3',
                 command=find_clas)
Button3.pack()

downLabel = Label(window,
                   padx=10,
                   pady=10,
                   text="Вставить текст можно только на английской раскладке!"
                        "\nЕсли у вас что-то не вышло, звоните или пишите: "
                        "\n8 962 996-91-09"
                        "\n maxmavr2005@gmail.com",
                   font=("Noah Regular Italic", 10),
                   bg="AntiqueWhite1",
                   fg="AntiqueWhite4")
downLabel.pack()

window.mainloop()
